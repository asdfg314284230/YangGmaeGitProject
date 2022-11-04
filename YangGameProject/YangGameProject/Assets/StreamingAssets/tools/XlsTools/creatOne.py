import openpyxl
import os

from openpyxl.chartsheet import Chartsheet

from const import source_path
from const import target_path


def creatXml():
    xmls = []
    i = 0
    for dir in os.listdir(source_path):
        print(i, dir)
        xmls.append(dir)
        i += 1
    # print(os.listdir(source_path))
    inputA = int(input("输入需要解析的表名ID : "))
    print("Begin 单个生成对应表格序号的解析文件-------------------------------------------------------")
    creatXml_One(xmls[inputA])
    print("End 单个生成对应表格序号的解析文件-------------------------------------------------------")


def creatXml_List(xlsList):
    for xls in xlsList:
        creatXml_One(xls)


def creatXml_One(xlsFileName):
    child = os.path.join(source_path, xlsFileName)
    if(child.find('~') != -1):
        return
    if os.path.isfile(child):
        # 打开一个workbook
        workbook = openpyxl.load_workbook(child)
        # 拿到所有sheet
        sheets = workbook.sheetnames
        # print(sheets)
        fileContent = ''
        fileContent += '<?xml version="1.0" encoding="utf-8"?>'
        fileContent += '\n<meta>'
        for i in range(len(sheets)):
            # 针对每一个sheet做操作
            if(str(sheets[i]).find('#') != -1):
                continue
            sheet = workbook[sheets[i]]
            if not isinstance(sheet, Chartsheet):
                fileContent += '\n  <model name="' + \
                    sheets[i] + 'Cfg' + '" ' + 'desc="' + "配置信息" + '">'
                if sheet.max_column is not None:
                    for r in range(1, sheet.max_column + 1):
                        if(sheet.cell(row=1, column=r).value is None):
                            continue
                        fileContent += '\n    <item tag="' + str(r) + '" ' + 'name="' + str(sheet.cell(row=1, column=r).value) + '" ' + 'type="' + str(
                            sheet.cell(row=3, column=r).value) + '" ' + 'desc="' + str(sheet.cell(row=2, column=r).value) + '" />'
                    fileContent += '\n  </model>'
        fileContent += '\n</meta>'
        fo = open(target_path + xlsFileName.split(".")[0] + 'Cfg.xml', "wb")
        fo.write(fileContent.encode('utf-8'))
        fo.close()


if __name__ == '__main__':
    creatXml()

import os
import re
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

TEST_FLAG = False


class ExcelRefInfo(object):
    def __init__(self,
                 xlsx_dir_path,
                 baseWorkbookName,
                 baseWorksheetName,
                 baseColumnName,
                 refWorkbookName,
                 refWorksheetName,
                 refColumnName):
        self.baseColumnInfo = ExcelColumnInfo(xlsx_dir_path,
                                              baseWorkbookName, baseWorksheetName, baseColumnName)
        self.refColumnInfo = ExcelColumnInfo(xlsx_dir_path,
                                             refWorkbookName, refWorksheetName, refColumnName)

    def check(self):
        if self.baseColumnInfo.initFlag is False or self.refColumnInfo.initFlag is False:
            return

        print("读取基础信息:"+str(self.baseColumnInfo))
        print("读取引用信息:"+str(self.refColumnInfo))
        if('*' in self.refColumnInfo.baseInfo.columnType):
            print('引用字段类型不能是数组:'+str(self.refColumnInfo))
            return
        print("开始检查引用丢失配置")
        ref_list = []
        for k, v in self.refColumnInfo.idDict.items():
            ref_list.append(str(v).strip())

        isArray = ('*' in self.baseColumnInfo.baseInfo.columnType)
        for k, v in self.baseColumnInfo.idDict.items():
            if(isArray):
                # 数组
                baseDataList = str(v).split(';')
                for baseData in baseDataList:
                    baseData = baseData.strip()
                    if baseData not in ref_list:
                        print("id={0} data={1} MISSING".format(
                            str(k), baseData))
                    else:
                        if(TEST_FLAG):
                            print("id={0} data={1} CORRECT".format(
                                str(k), baseData))
            else:
                # 非数组
                baseData = str(v).strip()
                if baseData not in ref_list:
                    print("id={0} data={1} MISSING".format(str(k), baseData))
                else:
                    if(TEST_FLAG):
                        print("id={0} data={1} CORRECT".format(
                            str(k), baseData))

        print("检查结束")
        print("")


class ExcelBaseColumnInfo(object):
    def __init__(self,
                 xlsx_dir_path,
                 workbookName,
                 worksheetName,
                 columnName,
                 columnType) -> None:
        self.dirpath = xlsx_dir_path
        self.workbookName = workbookName
        self.worksheetName = worksheetName
        self.columnName = columnName
        self.columnType = ''
        pass


class ExcelColumnInfo(object):
    def __init__(self,
                 xlsx_dir_path,
                 workbookName,
                 worksheetName,
                 columnName):
        self.initFlag = False
        self.baseInfo = ExcelBaseColumnInfo(xlsx_dir_path,
                                            workbookName,
                                            worksheetName,
                                            columnName,
                                            '')
        self.idDict = {}
        self.cellList = []
        self.check()

    def check(self):
        workbookPath = os.path.join(self.baseInfo.dirpath, self.baseInfo.workbookName+'.xlsx')
        if os.path.exists(workbookPath) and os.path.isfile(workbookPath):
            workbook = openpyxl.load_workbook(
                workbookPath, data_only=True)
            if workbook is not None:
                if self.baseInfo.worksheetName in workbook.sheetnames:
                    worksheet = workbook[self.baseInfo.worksheetName]
                    if worksheet is not None and isinstance(worksheet, Worksheet):
                        if worksheet.max_row > 4:
                            columnIdx = -1
                            for i in range(1, worksheet.max_column+1):
                                tempColumnName = worksheet.cell(1, i).value
                                if tempColumnName is not None \
                                        and isinstance(tempColumnName, str) \
                                        and tempColumnName is not '':
                                    if tempColumnName == self.baseInfo.columnName:
                                        # 如果有字段名
                                        tempColumnType = worksheet.cell(
                                            3, i).value
                                        if tempColumnType is not None \
                                                and isinstance(tempColumnType, str) \
                                                and tempColumnType is not '':
                                            columnIdx = i
                                            self.baseInfo.columnType = str(
                                                tempColumnType)
                                            break
                                        else:
                                            print('字段类型为空:'+str(self))
                            if columnIdx != -1:
                                for row in range(4, worksheet.max_row+1):
                                    id = worksheet.cell(
                                        row, 1).value
                                    dataCell = worksheet.cell(
                                        row, columnIdx)
                                    data = dataCell.value
                                    if id is not None:
                                        if data is not None:
                                            self.idDict[id] = str(data)
                                            self.cellList.append(dataCell)
                                self.initFlag = True
                            else:
                                print("找不到列:"+str(self))
                    else:
                        print("sheet格式有误:"+str(self))
                else:
                    print("文件中没有sheet:"+str(self))
                workbook.close()
            else:
                print("无法打开xlsx文件:"+str(self))
        else:
            print("当前路径不存在文件:"+workbookPath+" "+str(self))

    def __str__(self) -> str:
        return self.baseInfo.workbookName+" "+self.baseInfo.worksheetName+" "+self.baseInfo.columnName+" "+self.baseInfo.columnType


if __name__ == '__main__':
    lines = []
    with open('excel_ref_check_conf.txt', encoding='utf8') as f:
        lines = f.readlines()

    xlsx_dir_path = ''
    if len(lines) > 1:
        xlsx_dir_path = lines[0].replace('\n', '')
        if os.path.exists(xlsx_dir_path) and os.path.isdir(xlsx_dir_path):
            for idx in range(1, len(lines)):
                ref_info_line = lines[idx]
                ref_info_line = ref_info_line.replace('\n', '')
                ref_info_line = ref_info_line.strip()
                ref_info_line = re.sub('\\s+', ' ', ref_info_line)
                if(ref_info_line.startswith("#")):
                    continue
                ref_info_list = ref_info_line.split(' ')
                print(ref_info_list)
                if len(ref_info_list) > 5:
                    ref_info = ExcelRefInfo(
                        xlsx_dir_path,
                        ref_info_list[0],
                        ref_info_list[1],
                        ref_info_list[2],
                        ref_info_list[3],
                        ref_info_list[4],
                        ref_info_list[5],
                    )
                    ref_info.check()
                    pass
        else:
            print("xlsx文件夹不存在:"+xlsx_dir_path)
    else:
        print('excel_ref_check_conf.txt内容有误')
